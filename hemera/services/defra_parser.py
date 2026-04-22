"""Parse official DEFRA/DESNZ emission factor workbooks.

Reads the flat-format xlsx files (activity-based, Level 2) and the EEIO
ODS file (spend-based, Level 4) into lists of dicts matching the
EmissionFactor model.

Sources:
  - Activity: https://www.gov.uk/government/publications/greenhouse-gas-reporting-conversion-factors-{year}
  - EEIO:     https://www.gov.uk/government/statistics/uks-carbon-footprint
"""

from pathlib import Path
import openpyxl
import pandas as pd


# Rows with these scopes are excluded (biogenic CO2, not GHG Protocol)
_EXCLUDED_SCOPES = {"Outside of Scopes", "END", None}

# Only keep total CO2e rows, not individual gas breakdowns
_KEEP_GHG_UNIT = "kg CO2e"

# Map scope strings from the workbook to integers
_SCOPE_MAP = {
    "Scope 1": 1,
    "Scope 2": 2,
    "Scope 3": 3,
}


def parse_activity_factors(file_path: str, year: int) -> list[dict]:
    """Parse a DEFRA flat-format xlsx file into emission factor dicts.

    Args:
        file_path: Path to the flat-format xlsx workbook.
        year: The factor year (e.g. 2024).

    Returns:
        List of dicts with keys matching EmissionFactor model fields.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb["Factors by Category"]

    factors = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        row_id, scope_str, l1, l2, l3, l4, col_text, uom, ghg_unit, value = row[:10]

        # Filter: only total kg CO2e, exclude biogenic/outside scopes
        if ghg_unit != _KEEP_GHG_UNIT:
            continue
        if scope_str in _EXCLUDED_SCOPES:
            continue
        if value is None or value == 0:
            continue

        scope = _SCOPE_MAP.get(scope_str)
        if scope is None:
            continue

        # Build subcategory from the hierarchy levels below Level 1
        sub_parts = [p for p in (l2, l3, l4) if p]
        subcategory = " > ".join(sub_parts) if sub_parts else None

        # Build keywords from category hierarchy
        keyword_parts = [p for p in (l1, l2, l3, l4) if p]
        keywords = ",".join(p.lower() for p in keyword_parts)

        factors.append({
            "source": "defra",
            "category": l1,
            "subcategory": subcategory,
            "scope": scope,
            "factor_value": float(value),
            "unit": f"kgCO2e/{uom}",
            "factor_type": "activity",
            "year": year,
            "region": "UK",
            "keywords": keywords,
        })

    wb.close()
    return factors


_SKIP_SHEETS = {
    "Introduction", "What's new", "Index", "Conversions",
    "Fuel properties", "Haul definition", "Outside of scopes",
}


def parse_full_set_factors(file_path: str, year: int) -> list[dict]:
    """Parse a DEFRA full-set xlsx file (multi-sheet) into emission factor dicts.

    Each sheet has metadata rows 1-6, guidance text, then a data table whose
    header contains one or more "kg CO2e" columns.  Some sheets have
    sub-category headers in the row *above* the main header (e.g. "Diesel",
    "Petrol" for Passenger vehicles).

    Args:
        file_path: Path to the full-set xlsx workbook.
        year: The factor year (e.g. 2025).

    Returns:
        List of dicts with keys matching EmissionFactor model fields,
        plus source_sheet, source_row, source_hierarchy.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    factors: list[dict] = []

    for sheet_name in wb.sheetnames:
        if sheet_name in _SKIP_SHEETS:
            continue

        ws = wb[sheet_name]

        # --- Extract scope from row 6 col B ---
        scope_str = None
        for i, row in enumerate(ws.iter_rows(min_row=6, max_row=6, values_only=True), 6):
            scope_str = row[1] if len(row) > 1 else None

        scope = _SCOPE_MAP.get(scope_str)
        if scope is None:
            # Sheet has no valid scope — skip it
            continue

        # --- Scan for the header row containing "kg CO2e" ---
        header_row_idx = None
        header_row_vals = None
        prev_row_vals = None  # row above header (may have sub-category labels)
        rows_cache = []
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            rows_cache.append(row)
            if header_row_idx is None:
                if any(v == "kg CO2e" for v in row if v is not None):
                    header_row_idx = i
                    header_row_vals = row
                    prev_row_vals = rows_cache[-2] if len(rows_cache) >= 2 else None
            # Don't break — we'll process data rows from the cache after

        if header_row_idx is None:
            # No data table found on this sheet
            continue

        # --- Identify "kg CO2e" column positions and their sub-category labels ---
        co2e_positions = []  # list of (col_index, sub_label_or_None)
        for col_idx, val in enumerate(header_row_vals):
            if val == "kg CO2e":
                sub_label = None
                if prev_row_vals and col_idx < len(prev_row_vals):
                    lbl = prev_row_vals[col_idx]
                    if lbl and isinstance(lbl, str) and lbl.strip():
                        sub_label = lbl.strip()
                co2e_positions.append((col_idx, sub_label))

        if not co2e_positions:
            continue

        # --- Identify the column indices for Activity, Type/Fuel, Unit, Year ---
        col_map = {}  # name -> col_index
        for col_idx, val in enumerate(header_row_vals):
            if val and isinstance(val, str):
                name = val.strip()
                # Take the first occurrence of each name
                if name in ("Activity",) and "activity" not in col_map:
                    col_map["activity"] = col_idx
                elif name in ("Fuel", "Type", "Waste type", "Material",
                              "Country", "Size") and "type" not in col_map:
                    col_map["type"] = col_idx
                elif name == "Unit" and "unit" not in col_map:
                    col_map["unit"] = col_idx
                elif name == "Year" and "year" not in col_map:
                    col_map["year"] = col_idx

        if "activity" not in col_map:
            continue

        # --- Parse data rows ---
        data_rows = rows_cache[header_row_idx:]  # rows after the header
        current_activity = None
        current_type = None

        for offset, row in enumerate(data_rows):
            excel_row = header_row_idx + 1 + offset  # 1-based excel row number

            # Get cell values with safe indexing
            def cell(key):
                idx = col_map.get(key)
                if idx is not None and idx < len(row):
                    return row[idx]
                return None

            activity_val = cell("activity")
            type_val = cell("type")
            unit_val = cell("unit")
            year_val = cell("year")

            # Carry forward Activity (L1) from last non-None row
            if activity_val is not None:
                current_activity = str(activity_val).strip()
            # Carry forward Type (L2) from last non-None row
            if type_val is not None:
                current_type = str(type_val).strip()

            if unit_val is None or current_activity is None:
                continue

            unit_str = str(unit_val).strip()
            try:
                factor_year = int(year_val) if year_val is not None else year
            except (ValueError, TypeError):
                factor_year = year

            # Emit one factor per "kg CO2e" column that has a value
            for co2e_col_idx, sub_label in co2e_positions:
                if co2e_col_idx >= len(row):
                    continue
                value = row[co2e_col_idx]
                if value is None:
                    continue
                try:
                    fval = float(value)
                except (ValueError, TypeError):
                    continue
                if fval <= 0:
                    continue

                # Build the category / subcategory
                category = current_activity
                sub_parts = []
                if current_type:
                    sub_parts.append(current_type)
                if sub_label:
                    sub_parts.append(sub_label)
                subcategory = " > ".join(sub_parts) if sub_parts else None

                # Build keywords
                keyword_parts = [current_activity]
                if current_type:
                    keyword_parts.append(current_type)
                if sub_label:
                    keyword_parts.append(sub_label)
                keywords = ",".join(p.lower() for p in keyword_parts)

                # Source hierarchy
                hierarchy = [current_activity]
                if current_type:
                    hierarchy.append(current_type)

                factors.append({
                    "source": "defra",
                    "category": category,
                    "subcategory": subcategory,
                    "scope": scope,
                    "factor_value": fval,
                    "unit": f"kgCO2e/{unit_str}",
                    "factor_type": "activity",
                    "year": factor_year,
                    "region": "UK",
                    "keywords": keywords,
                    "source_sheet": sheet_name,
                    "source_row": excel_row,
                    "source_hierarchy": hierarchy,
                })

    wb.close()
    return factors


def parse_eeio_factors(file_path: str, year: int) -> list[dict]:
    """Parse the DEFRA EEIO spend-based factors (kgCO2e per GBP by SIC code).

    Args:
        file_path: Path to the EEIO ODS file.
        year: The factor year (e.g. 2022).

    Returns:
        List of dicts with keys matching EmissionFactor model fields.
    """
    df = pd.read_excel(file_path, engine="odf", header=None)

    factors = []
    for _, row in df.iterrows():
        sic_code = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else None
        description = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else None
        ghg_value = row.iloc[2]

        # Skip header/empty rows
        if sic_code is None or description is None:
            continue
        try:
            factor_value = float(ghg_value)
        except (ValueError, TypeError):
            continue

        keywords = description.lower()

        factors.append({
            "source": "defra-eeio",
            "category": description,
            "subcategory": sic_code,
            "scope": 3,
            "factor_value": factor_value,
            "unit": "kgCO2e/GBP",
            "factor_type": "spend",
            "year": year,
            "region": "UK",
            "keywords": keywords,
        })

    return factors
